
import csv
import io
from io import StringIO, BytesIO
from typing import Iterable, Iterator, Dict, Any
import datetime as dt
from datetime import datetime, date

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

def detect_delimiter(sample: str) -> str:
    return ";" if ";" in sample and sample.count(";") >= sample.count(",") else ","

def _parse_due_date(value: str | None):
    if not value:
        return None
    value = value.strip()
    if not value:
        return None
    # формат YYYY-MM-DD
    return dt.datetime.strptime(value, "%Y-%m-%d").date()


def _to_date(value):
    if not value:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    s = str(value).strip()
    if not s:
        return None
    # пробуем YYYY-MM-DD
    try:
        return dt.datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None

def parse_problems_csv(data: bytes) -> Iterator[dict]:
    """
    Ожидаемый CSV-хедер:
    list_code,number,title,assignee,due_date
    """
    text = data.decode("utf-8-sig")
    f = io.StringIO(text)
    reader = csv.DictReader(f)
    for row in reader:
        if not row.get("list_code") or not row.get("number"):
            continue
        list_code = row["list_code"].strip()
        number = int(row["number"])
        title = (row.get("title") or "").strip()
        assignee_raw = (row.get("assignee") or "").strip()
        assignee = int(assignee_raw) if assignee_raw else None
        due_date = _parse_due_date(row.get("due_date"))

        yield {
            "list_code": list_code,
            "number": number,
            "title": title,
            "assignee": assignee,
            "due_date": due_date,
        }


def _parse_assignees(value: Any) -> list[int]:
    """
    Парсит колонку assignee, где может быть:
      - один ID:   123456789
      - несколько: 123456789, 777777777, 999999999

    Возвращает список int.
    """
    if value is None:
        return []
    s = str(value).strip()
    if not s:
        return []

    ids: list[int] = []
    for part in s.split(","):
        part = part.strip()
        if not part:
            continue
        try:
            ids.append(int(part))
        except ValueError:
            # если мусор — просто пропускаем этот кусок
            continue
    return ids


def _normalize_due_date(value: Any) -> str | None:
    """
    Приводим дату к строке 'YYYY-MM-DD' или None, если даты нет/некорректна.
    """
    if value is None:
        return None

    # Excel-дата как datetime / date
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    s = str(value).strip()
    if not s:
        return None

    # Пробуем несколько форматов, если не вышло — возвращаем как есть
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            d = datetime.strptime(s, fmt).date()
            return d.isoformat()
        except ValueError:
            continue

    # Если формат нестандартный, но тебе ок — можно вернуть как есть
    return s


def parse_problems_xlsx(data: bytes) -> Iterator[Dict[str, object]]:
    """
    Парсер XLSX-файла со списком проблем.

    Ожидаемые колонки в первой строке (регистр не важен):

        number     / id      / №   — номер задачи в списке (обязателен)
        title                 — текст/описание задачи
        assignee              — Telegram ID исполнителя/исполнителей (через запятую)
        due_date              — срок исполнения (желательно в формате YYYY-MM-DD)

    На выходе даёт dict:
        {
            "number": int,
            "title": str,
            "assignees": list[int],
            "due_date": str | None,
        }
    """
    wb = load_workbook(BytesIO(data), data_only=True)
    ws = wb.active

    # --- читаем заголовок ---
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers: dict[str, int] = {}

    for idx, col_name in enumerate(header_row):
        if col_name is None:
            continue
        name = str(col_name).strip().lower()
        headers[name] = idx

    def get_col(row_values: tuple, *names: str) -> Any:
        """
        Берём значение по одному из возможных имён колонки.
        Например: get_col(row, "number", "id", "№")
        """
        for n in names:
            idx = headers.get(n)
            if idx is not None and idx < len(row_values):
                return row_values[idx]
        return None

    # --- пробегаем по строкам, начиная со 2-й ---
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue

        raw_number = get_col(row, "number", "id", "№")
        if raw_number is None:
            # строка без номера нам не интересна
            continue

        try:
            number = int(str(raw_number).strip())
        except ValueError:
            # некорректный номер — пропускаем
            continue

        title_val = get_col(row, "title")
        title = (str(title_val).strip() if title_val is not None else "") or f"Задача #{number}"

        assignee_val = get_col(row, "assignee")
        assignees = _parse_assignees(assignee_val)

        due_val = get_col(row, "due_date", "due", "deadline")
        due_date = _normalize_due_date(due_val)

        yield {
            "number": number,
            "title": title,
            "assignees": assignees,  # список int
            "due_date": due_date,    # строка 'YYYY-MM-DD' или None
        }
