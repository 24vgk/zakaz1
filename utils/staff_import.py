from io import BytesIO
from typing import List, Dict, Any

from openpyxl import load_workbook


def parse_staff_xlsx(data: bytes) -> List[Dict[str, Any]]:
    """
    Парсим zakaz.xlsx вида:
      assignee | post | fio

    Возвращает список словарей:
      {"assignee": int, "post": str | None, "fio": str | None}
    """
    wb = load_workbook(BytesIO(data), data_only=True)
    ws = wb.active

    # читаем шапку
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    cols: dict[str, int] = {}

    for idx, cell in enumerate(header_row, start=1):
        if cell.value is None:
            continue
        name = str(cell.value).strip().lower()
        cols[name] = idx

    # обязательный столбец — assignee
    if "assignee" not in cols:
        raise ValueError("В файле нет колонки 'assignee' (Telegram ID).")

    # необязательные
    post_col = cols.get("post")
    fio_col = cols.get("fio")

    rows: List[Dict[str, Any]] = []

    for row in ws.iter_rows(min_row=2):
        # assignee
        cell_ass = row[cols["assignee"] - 1].value
        if cell_ass is None:
            continue

        try:
            assignee = int(str(cell_ass).strip())
        except ValueError:
            # если ID кривой — пропускаем строку
            continue

        post = None
        fio = None

        if post_col is not None:
            v = row[post_col - 1].value
            post = str(v).strip() if v is not None else None

        if fio_col is not None:
            v = row[fio_col - 1].value
            fio = str(v).strip() if v is not None else None

        rows.append(
            {
                "assignee": assignee,
                "post": post,
                "fio": fio,
            }
        )

    return rows